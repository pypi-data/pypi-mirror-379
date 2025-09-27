"""Excel rendering core functionality."""

import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from markupsafe import Markup
from openpyxl.utils import get_column_letter

from .cache import ExcelCache
from .color_utils import get_rgb_from_color


class ExcelRenderer:
    """Core Excel rendering functionality."""

    def __init__(
        self,
        cache_enabled: bool = True,
        max_file_size_mb: int = 5,
        default_max_rows: int = 1000,
        default_max_cols: int = 50,
    ):
        self.cache_enabled = cache_enabled
        self.max_file_size_mb = max_file_size_mb
        self.default_max_rows = default_max_rows
        self.default_max_cols = default_max_cols
        self.cache = ExcelCache() if cache_enabled else None
        self.current_page = None

    def set_page_context(self, page):
        """Set the current page context for relative path resolution."""
        self.current_page = page

    def _resolve_file_path(self, file_path: str) -> str:
        """Resolve relative file paths based on current page context."""
        if os.path.isabs(file_path):
            return file_path

        if self.current_page and hasattr(self.current_page, "file"):
            page_dir = os.path.dirname(self.current_page.file.src_path)
            return os.path.join("docs", page_dir, file_path)

        return file_path

    def _check_file_exists(self, file_path: str) -> tuple[bool, str]:
        """Check if file exists and return status with resolved path."""
        resolved_path = self._resolve_file_path(file_path)
        exists = os.path.exists(resolved_path)
        return exists, resolved_path

    def _get_file_size_warning(self, file_path: str) -> str:
        """Get file size warning if file is too large."""
        try:
            file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB
            if file_size > self.max_file_size_mb:
                return f"""<div class='excel-warning'>
<p>⚠️ <strong>Large File Warning</strong></p>
<p>File size: {file_size:.1f}MB. This may affect page loading speed.</p>
<p>Consider splitting the data or using a smaller Excel file.</p>
</div>"""
        except OSError:
            pass
        return ""

    def render_excel_sheet(
        self,
        file_path: str,
        sheet_name: str,
        max_rows: int = None,
        max_cols: int = None,
    ) -> str:
        """Render a single Excel sheet as HTML table."""
        max_rows = max_rows or self.default_max_rows
        max_cols = max_cols or self.default_max_cols

        # Check file existence
        exists, resolved_path = self._check_file_exists(file_path)
        if not exists:
            return Markup(
                f"<div class='excel-error'><p>❌ <strong>Excel File Not Found</strong></p><p>File path: <code>{resolved_path}</code></p><p>Please check the file path and ensure the file exists.</p></div>"
            )

        # Performance warnings
        size_warning = self._get_file_size_warning(resolved_path)

        # Check cache
        cache_key = f"{resolved_path}#{sheet_name}"
        if self.cache and self.cache.is_valid(cache_key, resolved_path):
            return Markup(size_warning + self.cache.get(cache_key))

        try:
            workbook = openpyxl.load_workbook(resolved_path, data_only=True)

            if sheet_name not in workbook.sheetnames:
                available = "、".join(workbook.sheetnames)
                return Markup(
                    f"<div class='excel-error'><p>❌ <strong>Worksheet Not Found</strong></p><p>Requested sheet: <code>{sheet_name}</code></p><p>Available sheets: <code>{available}</code></p></div>"
                )

            sheet = workbook[sheet_name]

        except Exception as e:
            return Markup(
                f"<div class='excel-error'><p>❌ <strong>Excel File Read Error</strong></p><p>Error: <code>{str(e)}</code></p><p>Please check if the file is corrupted or in the correct format.</p></div>"
            )

        # Generate table HTML
        html = self._render_sheet_to_html(sheet, max_rows, max_cols)

        # Cache result
        if self.cache:
            self.cache.set(cache_key, html, resolved_path)

        return Markup(size_warning + html)

    def render_excel_all_sheets(
        self,
        file_path: str,
        max_rows: int = None,
        max_cols: int = None,
        include_sheets: List[str] = None,
        exclude_sheets: List[str] = None,
    ) -> str:
        """Render all sheets in an Excel file."""
        max_rows = max_rows or self.default_max_rows
        max_cols = max_cols or self.default_max_cols

        # Check file existence
        exists, resolved_path = self._check_file_exists(file_path)
        if not exists:
            return Markup(
                f"<div class='excel-error'><p>❌ <strong>Excel File Not Found</strong></p><p>File path: <code>{resolved_path}</code></p></div>"
            )

        try:
            workbook = openpyxl.load_workbook(resolved_path, data_only=True)
            all_sheets = workbook.sheetnames
        except Exception as e:
            return Markup(
                f"<div class='excel-error'><p>❌ <strong>Excel File Read Error</strong></p><p>Error: <code>{str(e)}</code></p></div>"
            )

        # Determine target sheets
        if include_sheets:
            target_sheets = [s for s in include_sheets if s in all_sheets]
            missing_sheets = [s for s in include_sheets if s not in all_sheets]
            if missing_sheets:
                return Markup(
                    f"<div class='excel-error'><p>❌ <strong>Specified Sheets Not Found</strong></p><p>Missing sheets: <code>{', '.join(missing_sheets)}</code></p><p>Available sheets: <code>{', '.join(all_sheets)}</code></p></div>"
                )
        else:
            exclude_sheets = exclude_sheets or []
            target_sheets = [s for s in all_sheets if s not in exclude_sheets]

        if not target_sheets:
            return Markup(
                "<div class='excel-warning'><p>⚠️ <strong>No Sheets to Display</strong></p><p>Please check include/exclude settings.</p></div>"
            )

        # Generate overview
        overview_html = f"<div class='excel-info'><p>📚 <strong>Excel File Overview</strong></p><p>File: <code>{os.path.basename(resolved_path)}</code></p><p>Displaying {len(target_sheets)} sheets: {', '.join(target_sheets)}</p></div>"

        # Render each sheet
        sheets_html = ""
        for i, sheet_name in enumerate(target_sheets):
            sheet_html = self.render_excel_sheet(
                file_path, sheet_name, max_rows, max_cols
            )
            sheets_html += f"<div class='excel-sheet-section'><h3 class='excel-sheet-title'>📊 Sheet: {sheet_name}</h3>{sheet_html}</div>"

            if i < len(target_sheets) - 1:
                sheets_html += "<hr class='excel-sheet-divider'>"

        return Markup(overview_html + sheets_html)

    def list_excel_sheets(self, file_path: str) -> str:
        """List all sheets in an Excel file."""
        exists, resolved_path = self._check_file_exists(file_path)
        if not exists:
            return Markup(
                f"<div class='excel-error'><p>❌ <strong>Excel File Not Found</strong></p><p>File path: <code>{resolved_path}</code></p></div>"
            )

        try:
            workbook = openpyxl.load_workbook(resolved_path, data_only=True)
            sheets = workbook.sheetnames

            sheet_list = ""
            for sheet in sheets:
                sheet_list += f"<li><code>{sheet}</code></li>"

            return Markup(
                f"<div class='excel-info'><p>📋 <strong>Sheet List</strong></p><p>File: <code>{os.path.basename(resolved_path)}</code></p><ol>{sheet_list}</ol></div>"
            )

        except Exception as e:
            return Markup(
                f"<div class='excel-error'><p>❌ Unable to read file: <code>{str(e)}</code></p></div>"
            )

    def _render_sheet_to_html(self, sheet, max_rows: int, max_cols: int) -> str:
        """Convert Excel sheet to HTML table."""
        # Get actual dimensions
        actual_rows = sheet.max_row or 0
        actual_cols = sheet.max_column or 0

        # Generate size info
        info_parts = [f"📊 Table size: {actual_rows} rows × {actual_cols} columns"]
        warning_parts = []

        # Check limits
        rows_truncated = actual_rows > max_rows
        cols_truncated = actual_cols > max_cols

        if rows_truncated:
            warning_parts.append(
                f"Row limit exceeded: showing first {max_rows} rows (total {actual_rows} rows)"
            )
        if cols_truncated:
            warning_parts.append(
                f"Column limit exceeded: showing first {max_cols} columns (total {actual_cols} columns)"
            )

        # Performance warning
        total_cells = min(actual_rows, max_rows) * min(actual_cols, max_cols)
        if total_cells > 10000:
            warning_parts.append(
                f"Large cell count ({total_cells:,} cells) may affect page performance"
            )

        # Build info HTML
        info_html = f"<div class='excel-info'><p>{' | '.join(info_parts)}</p></div>"

        warning_html = ""
        if warning_parts:
            warning_items = "".join([f"<li>{item}</li>" for item in warning_parts])
            warning_html = f"<div class='excel-warning'><p>⚠️ <strong>Display Limits</strong></p><ul>{warning_items}</ul><p><em>Tip: Adjust max_rows and max_cols parameters to control display range</em></p></div>"

        # Process merged cells
        merged_cells_map = {}
        for merged_range in sheet.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if (row, col) != (merged_range.min_row, merged_range.min_col):
                        merged_cells_map[(row, col)] = (
                            merged_range.min_row,
                            merged_range.min_col,
                        )

        # Generate table with wrapper
        html = "<div class='excel-table-wrapper'><table class='excel-table'>"
        process_rows = min(actual_rows, max_rows) if actual_rows else max_rows
        process_cols = min(actual_cols, max_cols) if actual_cols else max_cols

        for row_idx in range(1, process_rows + 1):
            html += "<tr>"
            for col_idx in range(1, process_cols + 1):
                if (row_idx, col_idx) in merged_cells_map:
                    continue

                cell = sheet.cell(row_idx, col_idx)
                html += self._render_cell_to_html(
                    cell, sheet, merged_cells_map, process_rows, process_cols
                )
            html += "</tr>"

        html += "</table></div>"

        return info_html + warning_html + html

    def _render_cell_to_html(
        self, cell, sheet, merged_cells_map, max_rows, max_cols
    ) -> str:
        """Render a single cell to HTML."""
        style_parts = []

        # Background color
        if cell.fill and cell.fill.start_color:
            try:
                bg_color = get_rgb_from_color(cell.fill.start_color)
                if bg_color:
                    style_parts.append(f"background-color: #{bg_color}")
            except:
                pass

        # Font styles
        if cell.font:
            # Font color
            if cell.font.color:
                try:
                    font_color = get_rgb_from_color(cell.font.color)
                    if font_color:
                        style_parts.append(f"color: #{font_color}")
                except:
                    pass

            if cell.font.bold:
                style_parts.append("font-weight: bold")
            if cell.font.italic:
                style_parts.append("font-style: italic")
            if cell.font.size:
                style_parts.append(f"font-size: {cell.font.size}pt")

        # Alignment
        if cell.alignment:
            if cell.alignment.horizontal:
                style_parts.append(f"text-align: {cell.alignment.horizontal}")
            if cell.alignment.vertical:
                valign_map = {"center": "middle", "top": "top", "bottom": "bottom"}
                if cell.alignment.vertical in valign_map:
                    style_parts.append(
                        f"vertical-align: {valign_map[cell.alignment.vertical]}"
                    )

        # Borders
        if cell.border:
            border_styles = []
            for side_name, side in [
                ("top", cell.border.top),
                ("right", cell.border.right),
                ("bottom", cell.border.bottom),
                ("left", cell.border.left),
            ]:
                if side.style:
                    width = (
                        "1px"
                        if side.style == "thin"
                        else "2px" if side.style == "thick" else "1px"
                    )
                    color = "#000000"
                    if side.color and side.color.rgb:
                        side_color = side.color.rgb
                        if isinstance(side_color, str) and len(side_color) == 8:
                            color = f"#{side_color[2:]}"
                    border_styles.append(f"border-{side_name}: {width} solid {color}")
            style_parts.extend(border_styles)

        # Handle merged cells
        colspan = rowspan = 1
        for merged_range in sheet.merged_cells.ranges:
            if (
                cell.row == merged_range.min_row
                and cell.column == merged_range.min_col
                and cell.row <= max_rows
                and cell.column <= max_cols
            ):
                colspan = min(merged_range.max_col, max_cols) - merged_range.min_col + 1
                rowspan = min(merged_range.max_row, max_rows) - merged_range.min_row + 1
                break

        style = "; ".join(style_parts)
        value = str(cell.value) if cell.value is not None else ""

        return (
            f"<td colspan='{colspan}' rowspan='{rowspan}' style='{style}'>{value}</td>"
        )
