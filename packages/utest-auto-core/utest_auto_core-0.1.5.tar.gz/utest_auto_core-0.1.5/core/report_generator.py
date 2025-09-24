"""
简化版报告生成模块

提供Excel和JSON格式的测试报告生成功能
包含测试结果、步骤详情、截图信息和全局监控结果
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
from dataclasses import asdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .test_case import TestResult, StepResult, TestStatus, StepStatus
from .config_manager import ReportConfig


logger = logging.getLogger(__name__)

# 简化版报告生成器，不再需要复杂的字体设置


class ReportGenerator:
    """简化版报告生成器"""
    
    def __init__(self, config: ReportConfig):
        """
        初始化报告生成器
        
        Args:
            config: 报告配置
        """
        self.config = config
        self.output_dir = Path(config.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # 简化版报告生成器，不需要创建子目录
    
    
    def generate_report(self, test_results: List[TestResult], 
                       device_info: Dict[str, Any] = None,
                       performance_data: Dict[str, Any] = None,
                       log_data: Dict[str, Any] = None,
                       global_monitor_result: Dict[str, Any] = None) -> Union[str, List[str]]:
        """
        生成测试报告
        
        Args:
            test_results: 测试结果列表
            device_info: 设备信息
            performance_data: 性能数据
            log_data: 日志数据
            global_monitor_result: 全局监控结果（ANR/Crash监控等）
            
        Returns:
            Union[str, List[str]]: 报告文件路径或路径列表
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 不再生成图表，简化报告
        
        # 根据配置生成相应格式的报告
        if self.config.report_format == "excel":
            return self._generate_excel_report(test_results, device_info, performance_data, log_data, timestamp, global_monitor_result)
        elif self.config.report_format == "json":
            return self._generate_json_report(test_results, device_info, performance_data, log_data, timestamp, global_monitor_result)
        elif self.config.report_format == "all":
            # 生成所有格式的报告
            reports = []
            reports.append(self._generate_excel_report(test_results, device_info, performance_data, log_data, timestamp, global_monitor_result))
            reports.append(self._generate_json_report(test_results, device_info, performance_data, log_data, timestamp, global_monitor_result))
            return reports
        else:
            raise ValueError(f"不支持的报告格式: {self.config.report_format}，支持的格式: excel, json, all")
    
    
    
    
    
    def _generate_excel_report(self, test_results: List[TestResult], 
                              device_info: Dict[str, Any],
                              performance_data: Dict[str, Any],
                              log_data: Dict[str, Any],
                              timestamp: str,
                              global_monitor_result: Dict[str, Any] = None) -> str:
        """
        生成简化版Excel报告
        
        Args:
            test_results: 测试结果列表
            device_info: 设备信息
            performance_data: 性能数据
            log_data: 日志数据
            timestamp: 时间戳
            global_monitor_result: 全局监控结果
            
        Returns:
            str: 报告文件路径
        """
        report_file = self.output_dir / f"test_report_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建测试概览工作表
        self._create_summary_sheet(wb, test_results, device_info, timestamp, global_monitor_result)
        
        # 创建测试详情工作表（包含步骤和截图信息）
        self._create_test_details_sheet(wb, test_results)
        
        # 创建性能指标汇总工作表
        self._create_performance_summary_sheet(wb, test_results)
        
        # 保存工作簿
        wb.save(report_file)
        logger.info(f"增强版Excel报告已生成: {report_file}")
        
        return str(report_file)
    
    
    
    
    def _create_summary_sheet(self, wb: Workbook, test_results: List[TestResult], 
                             device_info: Dict[str, Any], timestamp: str, 
                             global_monitor_result: Dict[str, Any] = None) -> None:
        """创建测试概览工作表"""
        ws = wb.create_sheet("测试概览", 0)
        
        # 设置标题样式
        title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置表头样式
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置数据样式
        data_font = Font(name="微软雅黑", size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:H1')
        ws['A1'] = f"自动化测试报告 - {timestamp}"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_alignment
        
        # 设备信息
        row = 3
        ws[f'A{row}'] = "设备信息"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = header_alignment
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        # 初始化设备信息数据列表
        device_info_data = []
        
        if device_info:
            device_info_data = [
                ("设备UDID", device_info.get('udid', 'N/A')),
                ("平台类型", device_info.get('platform', 'N/A')),
                ("设备型号", device_info.get('model', 'N/A')),
                ("制造商", device_info.get('manufacturer', 'N/A')),
                ("系统版本", device_info.get('os_version', 'N/A')),
                ("屏幕尺寸", f"{device_info.get('screen_width', 0)}x{device_info.get('screen_height', 0)}"),
                ("测试时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ]
            
            for i, (key, value) in enumerate(device_info_data):
                ws[f'A{row + i}'] = key
                ws[f'B{row + i}'] = str(value)
                ws[f'A{row + i}'].font = data_font
                ws[f'B{row + i}'].font = data_font
                ws[f'A{row + i}'].alignment = data_alignment
                ws[f'B{row + i}'].alignment = data_alignment
                ws[f'A{row + i}'].border = thin_border
                ws[f'B{row + i}'].border = thin_border
        else:
            # 当没有设备信息时，显示默认信息
            device_info_data = [
                ("设备信息", "未提供设备信息"),
                ("测试时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ]
            
            for i, (key, value) in enumerate(device_info_data):
                ws[f'A{row + i}'] = key
                ws[f'B{row + i}'] = str(value)
                ws[f'A{row + i}'].font = data_font
                ws[f'B{row + i}'].font = data_font
                ws[f'A{row + i}'].alignment = data_alignment
                ws[f'B{row + i}'].alignment = data_alignment
                ws[f'A{row + i}'].border = thin_border
                ws[f'B{row + i}'].border = thin_border
        
        # 全局监控结果
        if global_monitor_result:
            row += len(device_info_data) + 2
            ws[f'A{row}'] = "全局监控结果"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = header_alignment
            ws.merge_cells(f'A{row}:H{row}')
            
            row += 1
            monitor_data = [
                ("监控状态", "成功" if global_monitor_result.get('success', False) else "失败"),
                ("运行时间", f"{global_monitor_result.get('run_time', 0):.2f}秒"),
                ("Crash次数", global_monitor_result.get('crash_count', 0)),
                ("ANR次数", global_monitor_result.get('anr_count', 0)),
            ]
            
            for i, (key, value) in enumerate(monitor_data):
                ws[f'A{row + i}'] = key
                ws[f'B{row + i}'] = str(value)
                ws[f'A{row + i}'].font = data_font
                ws[f'B{row + i}'].font = data_font
                ws[f'A{row + i}'].alignment = data_alignment
                ws[f'B{row + i}'].alignment = data_alignment
                ws[f'A{row + i}'].border = thin_border
                ws[f'B{row + i}'].border = thin_border
            
            row += len(monitor_data) + 2
        
        # 测试统计
        ws[f'A{row}'] = "测试统计"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = header_alignment
        ws.merge_cells(f'A{row}:H{row}')
        
        # 统计测试结果
        total_tests = len(test_results)
        passed_tests = sum(1 for r in test_results if r.status == TestStatus.PASSED)
        failed_tests = sum(1 for r in test_results if r.status == TestStatus.FAILED)
        error_tests = sum(1 for r in test_results if r.status == TestStatus.ERROR)
        skipped_tests = sum(1 for r in test_results if r.status == TestStatus.SKIPPED)
        
        total_duration = sum(r.duration or 0 for r in test_results)
        
        row += 1
        stats_data = [
            ("总测试数", total_tests),
            ("通过", passed_tests),
            ("失败", failed_tests),
            ("错误", error_tests),
            ("跳过", skipped_tests),
            ("通过率", f"{(passed_tests/total_tests*100):.1f}%" if total_tests > 0 else "0%"),
            ("总耗时", f"{total_duration:.2f}秒")
        ]
        
        for i, (key, value) in enumerate(stats_data):
            ws[f'A{row + i}'] = key
            ws[f'B{row + i}'] = str(value)
            ws[f'A{row + i}'].font = data_font
            ws[f'B{row + i}'].font = data_font
            ws[f'A{row + i}'].alignment = data_alignment
            ws[f'B{row + i}'].alignment = data_alignment
            ws[f'A{row + i}'].border = thin_border
            ws[f'B{row + i}'].border = thin_border
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        
        # 设置行高
        for i in range(1, row + len(stats_data) + 1):
            ws.row_dimensions[i].height = 25

    def _create_performance_summary_sheet(self, wb: Workbook, test_results: List[TestResult]) -> None:
        """创建性能指标汇总工作表"""
        ws = wb.create_sheet("性能指标汇总")
        
        # 设置表头样式
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置数据样式
        data_font = Font(name="微软雅黑", size=11)
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:L1')
        ws['A1'] = "性能监控数据汇总"
        ws['A1'].font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # 表头
        headers = [
            "测试名称", "CPU使用率(%)", "内存峰值(MB)", "平均FPS", "卡顿率(%)",
            "上传流量(KB)", "下载流量(KB)", "应用版本", "设备型号", "监控时长(秒)",
            "数据点数", "监控状态"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填充性能数据
        current_row = 4
        for result in test_results:
            # 获取性能数据
            perf_data = result.performance_data or {}
            
            # 基础信息
            ws.cell(row=current_row, column=1, value=result.test_name).font = data_font
            ws.cell(row=current_row, column=1).border = thin_border
            
            # 客户需要的核心性能指标
            cpu_usage = perf_data.get('cpu_usage_avg', 0.0)
            memory_peak = perf_data.get('memory_peak_mb', 0.0)
            fps_avg = perf_data.get('fps_avg', 0.0)
            stutter_rate = perf_data.get('stutter_rate_percent', 0.0)
            upload_traffic = perf_data.get('network_upload_total_kb', 0.0)
            download_traffic = perf_data.get('network_download_total_kb', 0.0)
            
            # 性能指标数据
            ws.cell(row=current_row, column=2, value=f"{cpu_usage:.2f}").font = data_font
            ws.cell(row=current_row, column=3, value=f"{memory_peak:.2f}").font = data_font
            ws.cell(row=current_row, column=4, value=f"{fps_avg:.2f}").font = data_font
            ws.cell(row=current_row, column=5, value=f"{stutter_rate:.2f}").font = data_font
            ws.cell(row=current_row, column=6, value=f"{upload_traffic:.2f}").font = data_font
            ws.cell(row=current_row, column=7, value=f"{download_traffic:.2f}").font = data_font
            
            # 附加信息
            ws.cell(row=current_row, column=8, value=perf_data.get('app_version', 'N/A')).font = data_font
            ws.cell(row=current_row, column=9, value=perf_data.get('device_model', 'N/A')).font = data_font
            ws.cell(row=current_row, column=10, value=f"{perf_data.get('duration_sec', 0):.2f}").font = data_font
            ws.cell(row=current_row, column=11, value=perf_data.get('data_count', 0)).font = data_font
            
            # 监控状态
            monitor_status = "已启用" if perf_data else "未启用"
            ws.cell(row=current_row, column=12, value=monitor_status).font = data_font
            
            # 设置所有单元格的边框和对齐
            for col in range(1, 13):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = data_alignment
            
            # 根据性能指标设置颜色（可选）
            self._apply_performance_colors(ws, current_row, cpu_usage, fps_avg, stutter_rate)
            
            current_row += 1
        
        # 添加汇总统计行
        if test_results:
            self._add_performance_statistics(ws, test_results, current_row, data_font, data_alignment, thin_border)
        
        # 调整列宽
        column_widths = [25, 12, 12, 10, 10, 12, 12, 15, 20, 12, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 设置行高
        for row in range(1, current_row + 10):
            ws.row_dimensions[row].height = 25

    def _apply_performance_colors(self, ws, row: int, cpu_usage: float, fps_avg: float, stutter_rate: float) -> None:
        """根据性能指标设置单元格颜色"""
        try:
            # CPU使用率颜色（绿色：<50%, 黄色：50-80%, 红色：>80%）
            cpu_cell = ws.cell(row=row, column=2)
            if cpu_usage > 80:
                cpu_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif cpu_usage > 50:
                cpu_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            else:
                cpu_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            # FPS颜色（绿色：>30, 黄色：20-30, 红色：<20）
            fps_cell = ws.cell(row=row, column=4)
            if fps_avg < 20:
                fps_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif fps_avg < 30:
                fps_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            else:
                fps_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            # 卡顿率颜色（绿色：<5%, 黄色：5-10%, 红色：>10%）
            stutter_cell = ws.cell(row=row, column=5)
            if stutter_rate > 10:
                stutter_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif stutter_rate > 5:
                stutter_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            else:
                stutter_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
        except Exception as e:
            logger.warning(f"设置性能指标颜色失败: {e}")

    def _add_performance_statistics(self, ws, test_results: List[TestResult], start_row: int, 
                                  data_font: Font, data_alignment: Alignment, thin_border: Border) -> None:
        """添加性能统计汇总"""
        # 空行
        start_row += 1
        
        # 统计标题
        ws.merge_cells(f'A{start_row}:L{start_row}')
        ws[f'A{start_row}'] = "性能统计汇总"
        ws[f'A{start_row}'].font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
        ws[f'A{start_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws[f'A{start_row}'].alignment = Alignment(horizontal="center", vertical="center")
        
        start_row += 1
        
        # 收集所有性能数据
        all_cpu = []
        all_memory = []
        all_fps = []
        all_stutter = []
        all_upload = []
        all_download = []
        
        for result in test_results:
            perf_data = result.performance_data
            if perf_data:
                # 使用正确的字段名
                all_cpu.append(perf_data.get('cpu_usage_avg', 0))
                all_memory.append(perf_data.get('memory_peak_mb', 0))
                all_fps.append(perf_data.get('fps_avg', 0))
                all_stutter.append(perf_data.get('stutter_rate_percent', 0))
                all_upload.append(perf_data.get('network_upload_total_kb', 0))
                all_download.append(perf_data.get('network_download_total_kb', 0))
        
        # 计算统计值
        if all_cpu:
            stats_data = [
                ("指标", "平均值", "最大值", "最小值"),
                ("CPU使用率(%)", f"{sum(all_cpu)/len(all_cpu):.2f}", f"{max(all_cpu):.2f}", f"{min(all_cpu):.2f}"),
                ("内存峰值(MB)", f"{sum(all_memory)/len(all_memory):.2f}", f"{max(all_memory):.2f}", f"{min(all_memory):.2f}"),
                ("平均FPS", f"{sum(all_fps)/len(all_fps):.2f}", f"{max(all_fps):.2f}", f"{min(all_fps):.2f}"),
                ("卡顿率(%)", f"{sum(all_stutter)/len(all_stutter):.2f}", f"{max(all_stutter):.2f}", f"{min(all_stutter):.2f}"),
                ("上传流量(KB)", f"{sum(all_upload)/len(all_upload):.2f}", f"{max(all_upload):.2f}", f"{min(all_upload):.2f}"),
                ("下载流量(KB)", f"{sum(all_download)/len(all_download):.2f}", f"{max(all_download):.2f}", f"{min(all_download):.2f}")
            ]
            
            for i, row_data in enumerate(stats_data):
                for j, value in enumerate(row_data):
                    cell = ws.cell(row=start_row + i, column=j + 1, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    
                    # 设置表头样式
                    if i == 0:
                        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            # 没有性能数据
            ws.merge_cells(f'A{start_row}:D{start_row}')
            ws[f'A{start_row}'] = "无性能监控数据"
            ws[f'A{start_row}'].font = data_font
            ws[f'A{start_row}'].alignment = data_alignment
            ws[f'A{start_row}'].border = thin_border

    def _generate_performance_summary_json(self, test_results: List[TestResult]) -> Dict[str, Any]:
        """生成性能指标汇总JSON数据"""
        summary = {
            "total_tests": len(test_results),
            "tests_with_performance": 0,
            "performance_metrics": {
                "cpu_usage": {"avg": 0.0, "max": 0.0, "min": 0.0},
                "memory_peak": {"avg": 0.0, "max": 0.0, "min": 0.0},
                "fps": {"avg": 0.0, "max": 0.0, "min": 0.0},
                "stutter_rate": {"avg": 0.0, "max": 0.0, "min": 0.0},
                "network_upload": {"avg": 0.0, "max": 0.0, "min": 0.0},
                "network_download": {"avg": 0.0, "max": 0.0, "min": 0.0}
            },
            "test_details": []
        }
        
        # 收集所有性能数据
        all_cpu = []
        all_memory = []
        all_fps = []
        all_stutter = []
        all_upload = []
        all_download = []
        
        for result in test_results:
            perf_data = result.performance_data
            if perf_data:
                summary["tests_with_performance"] += 1
                
                # 提取客户需要的核心指标
                cpu_usage = perf_data.get('cpu_usage_avg', 0.0)
                memory_peak = perf_data.get('memory_peak_mb', 0.0)
                fps_avg = perf_data.get('fps_avg', 0.0)
                stutter_rate = perf_data.get('stutter_rate_percent', 0.0)
                upload_traffic = perf_data.get('network_upload_total_kb', 0.0)
                download_traffic = perf_data.get('network_download_total_kb', 0.0)
                
                # 添加到统计列表
                all_cpu.append(cpu_usage)
                all_memory.append(memory_peak)
                all_fps.append(fps_avg)
                all_stutter.append(stutter_rate)
                all_upload.append(upload_traffic)
                all_download.append(download_traffic)
                
                # 添加到测试详情
                test_detail = {
                    "test_name": result.test_name,
                    "status": result.status.value,
                    "performance_metrics": {
                        "cpu_usage_avg": cpu_usage,
                        "memory_peak_mb": memory_peak,
                        "fps_avg": fps_avg,
                        "stutter_rate_percent": stutter_rate,
                        "network_upload_total_kb": upload_traffic,
                        "network_download_total_kb": download_traffic
                    },
                    "additional_info": {
                        "app_version": perf_data.get('app_version', 'N/A'),
                        "device_model": perf_data.get('device_model', 'N/A'),
                        "duration_sec": perf_data.get('duration_sec', 0),
                        "data_count": perf_data.get('data_count', 0)
                    }
                }
                summary["test_details"].append(test_detail)
        
        # 计算统计值
        if all_cpu:
            summary["performance_metrics"]["cpu_usage"] = {
                "avg": round(sum(all_cpu) / len(all_cpu), 2),
                "max": round(max(all_cpu), 2),
                "min": round(min(all_cpu), 2)
            }
            summary["performance_metrics"]["memory_peak"] = {
                "avg": round(sum(all_memory) / len(all_memory), 2),
                "max": round(max(all_memory), 2),
                "min": round(min(all_memory), 2)
            }
            summary["performance_metrics"]["fps"] = {
                "avg": round(sum(all_fps) / len(all_fps), 2),
                "max": round(max(all_fps), 2),
                "min": round(min(all_fps), 2)
            }
            summary["performance_metrics"]["stutter_rate"] = {
                "avg": round(sum(all_stutter) / len(all_stutter), 2),
                "max": round(max(all_stutter), 2),
                "min": round(min(all_stutter), 2)
            }
            summary["performance_metrics"]["network_upload"] = {
                "avg": round(sum(all_upload) / len(all_upload), 2),
                "max": round(max(all_upload), 2),
                "min": round(min(all_upload), 2)
            }
            summary["performance_metrics"]["network_download"] = {
                "avg": round(sum(all_download) / len(all_download), 2),
                "max": round(max(all_download), 2),
                "min": round(min(all_download), 2)
            }
        
        return summary
    
    def _create_test_details_sheet(self, wb: Workbook, test_results: List[TestResult]) -> None:
        """创建测试详情工作表（每个步骤独占一行）"""
        ws = wb.create_sheet("测试详情")
        
        # 设置表头
        headers = ["测试名称", "状态", "开始时间", "结束时间", "耗时(秒)", "步骤名称", "步骤状态", "步骤错误", "截图", "性能监控", "Logcat", "录制"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        current_row = 2
        for result in test_results:
            status_color = self._get_status_color(result.status)
            
            if result.steps:
                # 每个步骤一行
                start_row = current_row
                for step_idx, step in enumerate(result.steps):
                    # 测试用例基本信息（只在第一步时填写）
                    if step_idx == 0:
                        ws.cell(row=current_row, column=1, value=result.test_name)
                        ws.cell(row=current_row, column=2, value=result.status.value)
                        ws.cell(row=current_row, column=3, value=result.start_time.strftime("%Y-%m-%d %H:%M:%S") if result.start_time else "")
                        ws.cell(row=current_row, column=4, value=result.end_time.strftime("%Y-%m-%d %H:%M:%S") if result.end_time else "")
                        ws.cell(row=current_row, column=5, value=f"{result.duration:.2f}" if result.duration else "")
                        
                        # 设置状态颜色
                        status_cell = ws.cell(row=current_row, column=2)
                        status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                        status_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
                    
                    # 步骤信息
                    ws.cell(row=current_row, column=6, value=step.step_name)
                    ws.cell(row=current_row, column=7, value=step.status.value)
                    ws.cell(row=current_row, column=8, value=step.error_message or "")
                    
                    # 设置步骤状态颜色
                    step_status_cell = ws.cell(row=current_row, column=7)
                    step_status_color = self._get_status_color(step.status)
                    step_status_cell.fill = PatternFill(start_color=step_status_color, end_color=step_status_color, fill_type="solid")
                    step_status_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
                    
                    # 插入步骤截图
                    if step.screenshots:
                        self._insert_screenshots_to_cell(ws, current_row, 9, step.screenshots)
                    else:
                        ws.cell(row=current_row, column=9, value="无截图")
                    
                    # 监控数据（只在第一步时显示）
                    if step_idx == 0:
                        # 性能监控数据
                        perf_text = self._format_monitor_data(result.performance_data, "性能监控")
                        ws.cell(row=current_row, column=10, value=perf_text)
                        
                        # Logcat数据
                        logcat_text = self._format_monitor_data(result.logcat_data, "Logcat")
                        ws.cell(row=current_row, column=11, value=logcat_text)
                        
                        # 录制数据
                        recording_text = self._format_monitor_data(result.recording_data, "录制")
                        ws.cell(row=current_row, column=12, value=recording_text)
                    
                    current_row += 1
                
                # 合并测试用例基本信息的单元格
                if len(result.steps) > 1:
                    ws.merge_cells(f'A{start_row}:A{current_row - 1}')
                    ws.merge_cells(f'B{start_row}:B{current_row - 1}')
                    ws.merge_cells(f'C{start_row}:C{current_row - 1}')
                    ws.merge_cells(f'D{start_row}:D{current_row - 1}')
                    ws.merge_cells(f'E{start_row}:E{current_row - 1}')
                    ws.merge_cells(f'J{start_row}:J{current_row - 1}')  # 性能监控
                    ws.merge_cells(f'K{start_row}:K{current_row - 1}')  # Logcat
                    ws.merge_cells(f'L{start_row}:L{current_row - 1}')  # 录制
                    
                    # 设置合并单元格的对齐方式
                    for col in [1, 2, 3, 4, 5, 10, 11, 12]:
                        cell = ws.cell(row=start_row, column=col)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # 没有步骤的测试用例
                ws.cell(row=current_row, column=1, value=result.test_name)
                ws.cell(row=current_row, column=2, value=result.status.value)
                ws.cell(row=current_row, column=3, value=result.start_time.strftime("%Y-%m-%d %H:%M:%S") if result.start_time else "")
                ws.cell(row=current_row, column=4, value=result.end_time.strftime("%Y-%m-%d %H:%M:%S") if result.end_time else "")
                ws.cell(row=current_row, column=5, value=f"{result.duration:.2f}" if result.duration else "")
                ws.cell(row=current_row, column=6, value="无步骤")
                ws.cell(row=current_row, column=7, value="")
                ws.cell(row=current_row, column=8, value=result.error_message or "")
                ws.cell(row=current_row, column=9, value="无截图")
                
                # 监控数据
                perf_text = self._format_monitor_data(result.performance_data, "性能监控")
                ws.cell(row=current_row, column=10, value=perf_text)
                
                logcat_text = self._format_monitor_data(result.logcat_data, "Logcat")
                ws.cell(row=current_row, column=11, value=logcat_text)
                
                recording_text = self._format_monitor_data(result.recording_data, "录制")
                ws.cell(row=current_row, column=12, value=recording_text)
                
                # 设置状态颜色
                status_cell = ws.cell(row=current_row, column=2)
                status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                status_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
                
                current_row += 1
        
        # 调整列宽
        column_widths = [25, 10, 20, 20, 12, 30, 12, 30, 80, 30, 30, 30]  # 增加监控数据列宽度
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 设置行高
        for row in range(1, current_row):
            ws.row_dimensions[row].height = 150  # 增加行高以容纳更多截图
    
    def _get_status_color(self, status: Union[TestStatus, StepStatus]) -> str:
        """获取状态对应的颜色"""
        color_map = {
            TestStatus.PASSED: "00B050",
            TestStatus.FAILED: "FF0000",
            TestStatus.ERROR: "FF6600",
            TestStatus.SKIPPED: "FFC000",
            TestStatus.RUNNING: "0070C0",
            TestStatus.PENDING: "808080",
            StepStatus.PASSED: "00B050",
            StepStatus.FAILED: "FF0000",
            StepStatus.ERROR: "FF6600",
            StepStatus.SKIPPED: "FFC000",
            StepStatus.RUNNING: "0070C0",
            StepStatus.PENDING: "808080"
        }
        return color_map.get(status, "808080")
    
    def _generate_json_report(self, test_results: List[TestResult], 
                             device_info: Dict[str, Any],
                             performance_data: Dict[str, Any],
                             log_data: Dict[str, Any],
                             timestamp: str,
                             global_monitor_result: Dict[str, Any] = None) -> str:
        """生成JSON报告"""
        report_file = self.output_dir / f"test_report_{timestamp}.json"
        
        # 转换测试结果为字典
        results_data = []
        for result in test_results:
            result_dict = asdict(result)
            # 转换datetime对象为字符串
            if result_dict['start_time']:
                result_dict['start_time'] = result_dict['start_time'].isoformat()
            if result_dict['end_time']:
                result_dict['end_time'] = result_dict['end_time'].isoformat()
            
            # 转换步骤结果
            steps_data = []
            for step in result.steps:
                step_dict = asdict(step)
                if step_dict['start_time']:
                    step_dict['start_time'] = step_dict['start_time'].isoformat()
                if step_dict['end_time']:
                    step_dict['end_time'] = step_dict['end_time'].isoformat()
                steps_data.append(step_dict)
            result_dict['steps'] = steps_data
            
            results_data.append(result_dict)
        
        # 构建报告数据
        report_data = {
            "report_info": {
                "generated_at": datetime.now().isoformat(),
                "timestamp": timestamp,
                "format": "json"
            },
            "device_info": device_info,
            "test_results": results_data,
            "performance_data": performance_data,
            "log_data": log_data,
            "global_monitor_result": global_monitor_result,
            "performance_summary": self._generate_performance_summary_json(test_results)
        }
        
        # 保存JSON文件
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"JSON报告已生成: {report_file}")
        return str(report_file)
    
    def _insert_screenshots_to_cell(self, ws, row: int, col: int, screenshots: List[str]) -> None:
        """将截图插入到Excel单元格中"""
        try:
            from openpyxl.drawing import image as xl_image
            
            # 如果只有一个截图，直接插入
            if len(screenshots) == 1:
                screenshot_path = screenshots[0]
                if os.path.exists(screenshot_path):
                    img = xl_image.Image(screenshot_path)
                    # 调整图片大小以适应单元格
                    img.width = 200
                    img.height = 150
                    img.anchor = f"{get_column_letter(col)}{row}"
                    ws.add_image(img)
                else:
                    ws.cell(row=row, column=col, value=f"截图文件不存在: {screenshot_path}")
            else:
                # 多个截图，创建缩略图网格
                self._insert_multiple_screenshots(ws, row, col, screenshots)
                
        except Exception as e:
            logger.warning(f"插入截图失败: {e}")
            # 如果插入失败，显示文件路径
            screenshot_text = "\n".join([f"{i+1}. {screenshot}" for i, screenshot in enumerate(screenshots)])
            ws.cell(row=row, column=col, value=screenshot_text)
    
    def _format_monitor_data(self, monitor_data: Dict[str, Any], monitor_type: str) -> str:
        """格式化监控数据为可读文本"""
        if not monitor_data:
            return f"{monitor_type}: 未启用"
        
        # 构建信息
        info_lines = [f"{monitor_type}: 已启用"]
        
        # 添加文件路径
        if 'file_path' in monitor_data:
            info_lines.append(f"文件路径: {monitor_data['file_path']}")
        elif 'log_file' in monitor_data:
            info_lines.append(f"日志文件: {monitor_data['log_file']}")
        elif 'video_file' in monitor_data:
            info_lines.append(f"视频文件: {monitor_data['video_file']}")
        
        # 添加文件大小（如果存在）
        if 'file_size' in monitor_data:
            size_mb = monitor_data['file_size'] / (1024 * 1024)
            info_lines.append(f"文件大小: {size_mb:.2f}MB")
        
        # 性能监控特殊处理
        if monitor_type == "性能监控":
            if 'app_version' in monitor_data:
                info_lines.append(f"应用版本: {monitor_data['app_version']}")
            if 'device_model' in monitor_data:
                info_lines.append(f"设备型号: {monitor_data['device_model']}")
            if 'data_count' in monitor_data:
                info_lines.append(f"数据点数: {monitor_data['data_count']}")
            
            # 客户需要的核心性能指标
            if 'cpu_usage_avg' in monitor_data:
                info_lines.append(f"CPU使用率: {monitor_data['cpu_usage_avg']:.2f}%")
            if 'memory_peak_mb' in monitor_data:
                info_lines.append(f"内存峰值: {monitor_data['memory_peak_mb']:.2f}MB")
            if 'fps_avg' in monitor_data:
                info_lines.append(f"平均FPS: {monitor_data['fps_avg']:.2f}")
            if 'stutter_rate_percent' in monitor_data:
                info_lines.append(f"卡顿率: {monitor_data['stutter_rate_percent']:.2f}%")
            if 'network_upload_total_kb' in monitor_data:
                info_lines.append(f"上传流量: {monitor_data['network_upload_total_kb']:.2f}KB")
            if 'network_download_total_kb' in monitor_data:
                info_lines.append(f"下载流量: {monitor_data['network_download_total_kb']:.2f}KB")
            
            # 兼容旧的字段名
            if 'avg_fps' in monitor_data and 'fps_avg' not in monitor_data:
                info_lines.append(f"平均FPS: {monitor_data['avg_fps']:.1f}")
            if 'avg_cpu_usage' in monitor_data and 'cpu_usage_avg' not in monitor_data:
                info_lines.append(f"平均CPU: {monitor_data['avg_cpu_usage']:.1f}%")
            if 'avg_memory_usage' in monitor_data and 'memory_peak_mb' not in monitor_data:
                info_lines.append(f"平均内存: {monitor_data['avg_memory_usage']:.1f}MB")
        
        # 添加持续时间
        if 'duration' in monitor_data and monitor_data['duration'] > 0:
            info_lines.append(f"持续时间: {monitor_data['duration']:.2f}秒")
        
        return "\n".join(info_lines)
    
    def _insert_multiple_screenshots(self, ws, row: int, col: int, screenshots: List[str]) -> None:
        """插入多个截图到单元格中"""
        try:
            from openpyxl.drawing import image as xl_image
            from PIL import Image as PILImage
            import io
            
            # 创建缩略图网格
            max_images_per_row = 2  # 每行最多2个图片
            thumbnail_size = (100, 75)  # 缩略图大小
            
            # 计算网格布局
            total_images = min(len(screenshots), 4)  # 最多显示4个截图
            rows = (total_images + max_images_per_row - 1) // max_images_per_row
            
            for i, screenshot_path in enumerate(screenshots[:total_images]):
                if not os.path.exists(screenshot_path):
                    continue
                    
                # 计算位置
                img_row = i // max_images_per_row
                img_col = i % max_images_per_row
                
                # 创建缩略图
                try:
                    with PILImage.open(screenshot_path) as img:
                        img.thumbnail(thumbnail_size, PILImage.Resampling.LANCZOS)
                        
                        # 保存到内存
                        img_buffer = io.BytesIO()
                        img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        # 插入到Excel
                        xl_img = xl_image.Image(img_buffer)
                        xl_img.width = thumbnail_size[0]
                        xl_img.height = thumbnail_size[1]
                        
                        # 计算锚点位置
                        anchor_col = col + img_col
                        anchor_row = row + img_row
                        xl_img.anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
                        
                        ws.add_image(xl_img)
                        
                except Exception as e:
                    logger.warning(f"处理截图 {screenshot_path} 失败: {e}")
                    continue
            
            # 如果有更多截图，显示提示
            if len(screenshots) > total_images:
                ws.cell(row=row + rows, column=col, value=f"... 还有 {len(screenshots) - total_images} 个截图")
                
        except Exception as e:
            logger.warning(f"插入多个截图失败: {e}")
            # 如果失败，显示文件路径列表
            screenshot_text = "\n".join([f"{i+1}. {screenshot}" for i, screenshot in enumerate(screenshots)])
            ws.cell(row=row, column=col, value=screenshot_text)
