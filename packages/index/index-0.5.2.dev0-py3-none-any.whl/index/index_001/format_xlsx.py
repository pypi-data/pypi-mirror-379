#!/usr/bin/env python
# coding=utf-8
# Stan 2024-11-01

from openpyxl import load_workbook

from ..chunk import chunk
from ..timer import Timer


def main_yield(filename, db, options):
    with Timer(f"[ {__name__} ] load_workbook", db.verbose) as t:
        book = load_workbook(filename, read_only=True, data_only=True)

    sheet_names = book.get_sheet_names()
    sheet_list  = options.get('sheets', sheet_names)
    chunk_rows  = options.get('chunk_rows', 5000)

    for name in sheet_list:             # 1-based integer or string
        # 1-based integer and string
        shid, shname = get_shid_name(sheet_names, name)
        if shid is None:
            db.push_task_record('warning', f"Wrong sheed name/id: {name}")

        sh = book[shname]               # string
#       sh = book.worksheets[shid-1]    # worksheets accepts 0-based

        if db.verbose:
            print(f"Processing: # {shid} ({sh.title}) / max_row: {sh.max_row}, max_column: {sh.max_column}")

        for ki, chunk_i in enumerate(chunk(sh.iter_rows(), chunk_rows)):
            records = []

            for kj, row in enumerate(chunk_i):
                row_values = get_row_values(row)
                if row_values:
                    idx = ki * chunk_rows + kj
                    _r = idx + 1

                    record = dict(row=row_values, _r=_r)
                    records.append(record)

            yield records, {
                '_shid': shid,
#               '_shname': sh.title
            }
            records = []        # release memory

    book.close()


def get_shid_name(sheet_names, name):
    if isinstance(name, int):
        if len(sheet_names) < name:
            return None, None

        shid0 = name - 1
        name = sheet_names[shid0]

    else:
        if name not in sheet_names:
            return None, None

        shid0 = sheet_names.index(name)

    return shid0 + 1, name


def get_row_values(row):
#   values = [get_strip(i.value) for i in row]
    values = [get_strip(f"ERROR({i.value})" \
              if i.data_type == 'e' else i.value) for i in row]
    if any(x is not None for x in values):
        return values

    return []


def get_strip(s):
    if isinstance(s, str):
        return s.strip()

    return s
