#!/usr/bin/env python
# coding=utf-8
# Stan 2024-11-01

from openpyxl import load_workbook

# from ..chunk import chunk
from ..timer import Timer


def main_yield(filename, config):
    verbose = config.get('verbose')

    with Timer(f"[ {__name__} ] load_workbook '{filename}'", verbose) as t:
        book = load_workbook(filename, read_only=True, data_only=True)

    sheet_list  = config.get('sheets', book.get_sheet_names())
    chunk_rows  = config.get('chunk_rows', 5000)

    for index in sheet_list:            # integer or string
        if isinstance(index, int):      # index is 1-based, worksheets accepts 0-based
            sh = book.worksheets[index-1]
        else:
            sh = book[index]            # index is a string

        if verbose:
            print(f"Processing: # {index} ({sh.title}) / max_row: {sh.max_row}, max_column: {sh.max_column}")

#       for ki, chunk_i in enumerate(chunk(sh.rows, chunk_rows)):   # chunk здесь не работает
        for chunk_i in [sh.rows]:
            records = []

            for kj, row in enumerate(chunk_i):
                idx = kj
                _r = idx + 1

                row_values = [i.value for i in row]
                if any(x is not None for x in row_values):
                    record = dict(row=row_values, _r=_r)
                    records.append(record)

            yield records, {
                '_shid': index,
#               '_shname': sh.title
            }
            records = []        # release memory

    book.close()
