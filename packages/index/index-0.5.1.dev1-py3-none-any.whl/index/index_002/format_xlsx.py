#!/usr/bin/env python
# coding=utf-8
# Stan 2024-11-01

from openpyxl import load_workbook

# from ..chunk import chunk
from ..timer import Timer


def main_yield(filename, db, config, f_id, **kargs):
    verbose = kargs.get('verbose')

    with Timer(f"[ {__name__} ] load_workbook '{filename}'", verbose) as t:
        book = load_workbook(filename, read_only=True, data_only=True)

    nrows       = config['nrows']
    rows_values = config['rows_values']
    keys        = config['keys']

    sheet_list  = config.get('sheets', book.get_sheet_names())
    chunk_rows  = config.get('chunk_rows', 5000)

    for index in sheet_list:            # integer or string
        if isinstance(index, int):      # index is 1-based, worksheets accepts 0-based
            sh = book.worksheets[index-1]
        else:
            sh = book[index]            # index is a string

        if verbose:
            print(f"Processing: # {index} ({sh.title}) / max_row: {sh.max_row}, max_column: {sh.max_column}")

        res = find_table_header(sh, nrows, rows_values)
        if not res:
            raise Exception("Table header not found")

        header_row, data_row, indices = res

        # Массивы внутри indices должны содержать ровно по одному элементу
        # Если элементов нет - данная колонка не была найдена
        # Если элементов > 1 - данная колонка встречается в таблице дважды
        unpredictable_indexes = list(filter(lambda x: len(x) != 1, indices))
        if unpredictable_indexes:
            db.push_file_record(
                f_id,
                "find_table_header",
                shname = sh.title,
                error  = true,
                msg    = "unpredictable_indexes",
                res    = unpredictable_indexes
            )
            raise Exception(f"unpredictable_indexes: {unpredictable_indexes}")

        indices = [i[0] for i in indices]
        db.push_file_record(
            f_id,
            "find_table_header",
            shname     = sh.title,
            header_row = header_row,
            data_row   = data_row,
            indices    = indices,
            max_row    = sh.max_row,
            max_column = sh.max_column
        )

#       for ki, chunk_i in enumerate(chunk(sh.rows, chunk_rows)):   # chunk здесь не работает
        for chunk_i in [sh.rows]:
            records = []

            for kj, row in enumerate(chunk_i):
                idx = kj
                _r = idx + 1

                if idx < data_row:
                    continue

                row_values = [i.value for i in (map(row.__getitem__, indices))]
                record = dict(zip(keys, row_values), _r=_r)
                records.append(record)

            yield records, {
                '_shid': index,
#               '_shname': sh.title
            }
            records = []


def find_table_header(sh, nrows, rows_values):
    header_row = None
    data_row = None

    current_row = 0
    current_col_values = [i[current_row] for i in rows_values]

    all_indices = []

    for row in sh.rows:
#       success = check_entries(current_col_values, [i.v for i in row])
        success, indices = find_entries(current_col_values, [i.value for i in row])

        if success:
            if header_row is None:
                header_row = current_row
                all_indices = indices

            else:
                all_indices = filter_indices(all_indices, indices)

            current_row += 1

            if current_row == nrows:
                data_row = current_row

                return header_row, data_row, all_indices

            current_col_values = [i[current_row] for i in rows_values]

    return None


def check_entries(row1, row2):
    for v in row1:
        res = [v in row2 for v in row1]

    fault = False in res
    return not fault


def find_entries(row1, row2):
    indices = []
    for v in row1:
        indices.append([i for i, x in enumerate(row2) if x == v])

    fault = [] in indices
    return not fault, indices


def filter_indices(indices1, indices2):
    accumulate = []
    for i in range(len(indices1)):
        accumulate.append( list(set(indices1[i]) & set(indices2[i])) )

    return accumulate
