#!/usr/bin/env python
"""
Process the ECCC oil records that have both types of distillation sets.
( /gnome/oil_database/oil_database/-/issues/587 )

There should be 106 such records total in the source data.
87 records were added as new oils and are presumed to be renamed as AD records
per Rin's list of similar oils.  So we will probably need to go off of
that list for the reconciled distillation data to make it to its
"final destination".

I have placed the spreadsheet in the same folder as this script
("./CC_to_AD.xlsx").

Input: a list of CC related oil entries that were kept.

    - Column 1 ("CC ID"): The CC ID oil entry that is still up on the
                          Stage Server.
    - Column 2 ("New AD ID"): The AD ID that was a match to the CC in the
                              first column.

- We need to get a mapping of the AD ID and the source ID
"""
import os
import sys
from pathlib import Path
from argparse import ArgumentParser

import csv

from openpyxl import load_workbook

import adios_db
from adios_db.models.oil.oil import Oil

from pprint import pprint


argp = ArgumentParser(description='Script Arguments:')
argp.add_argument('--path', nargs=1,
                  help=('Specify a path to a data storage area (filesystem). '
                        'If not specified, the default is to use "./data"'))
argp.add_argument('-d', '--dry_run', action='store_true',
                  help=('Do not perform the file actions, only print '
                        'the actions out.'))

def generate_sheet(file):
    wb = load_workbook(file, data_only=True)
    return wb['Sheet1']  # First sheet is not set in the file, use default.


def generate_xls_row_iter(sheet):
    sheet_iter = sheet.iter_rows()
    sheet_iter.__next__()  # bypass column names

    return sheet_iter


def generate_csv_row_iter(csvfile):
    csv_reader = csv.reader(csvfile)
    csv_reader.__next__()  # bypass column names

    return csv_reader


def oil_json_file_path(base_path, oil_id):
    return (Path(base_path) / 'oil' / oil_id[:2] / f'{oil_id}.json')


def generate_source_id_lu(base_path):
    """
    Generate a lookup table for our ECCC source IDs that were renamed to
    ADIOS IDs in the noaa-oil-data repo.
    """
    lookup_dict = {}
    sheet = generate_sheet(Path('./CC_to_AD.xlsx'))

    for r in generate_xls_row_iter(sheet):
        cc_id, ad_id = [f.value for f in r]
        ad_file = oil_json_file_path(base_path, ad_id)

        if ad_file.is_file():
            ad_oil = Oil.from_file(ad_file)

            lookup_dict[ad_oil.metadata.source_id] = ad_oil.oil_id
        else:
            print(f'{ad_file} file is missing.')
    return lookup_dict


def reconcile_file_paths(base_path, id_lookup, src_id, cc_id, dry_run=False):
    ad_id = id_lookup.get(src_id, None)

    if ad_id is not None:
        cc_file = oil_json_file_path(base_path, cc_id)
        ad_file = oil_json_file_path(base_path, ad_id)
        
        if cc_file.is_file() and ad_file.is_file():
            print(f'{ad_file} update the file...')
            update_oil_file(cc_file, ad_file, dry_run)
        elif cc_file.is_file():
            print(f'{ad_file} destination file is missing...')
        else:
            print(f'{cc_file} source file is missing...')
    else:
        print(f'{src_id} no destination file...')


def update_oil_file(src_file, dest_file, dry_run=False):
    if dry_run:
        return

    src_oil = Oil.from_file(src_file)
    dest_oil = Oil.from_file(dest_file)

    for src_smpl, dest_smpl in zip(src_oil.sub_samples, dest_oil.sub_samples):
        src_dist = src_smpl.distillation_data
        dest_dist = dest_smpl.distillation_data

        dest_dist.cuts = src_dist.cuts

    dest_oil.to_file(dest_file)


def main(argv=sys.argv):
    args = argp.parse_args(argv[1:])

    base_path = Path(args.path[0]) if args.path is not None else Path('./data')
    print(f'Work directory: {base_path}')

    dry_run = args.dry_run
    if dry_run:
        print(f'Dry run...no files to be changed.')

    id_lookup = generate_source_id_lu(base_path)

    with open('./recs_with_both_dist_sets.csv', newline='') as csvfile:
        for r in generate_csv_row_iter(csvfile):
            src_id, cc_id, _, _, _ = [f.strip() for f in r]
            reconcile_file_paths(base_path, id_lookup, src_id, cc_id, dry_run)


if __name__ == "__main__":
    main()
