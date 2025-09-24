#!/usr/bin/env python
"""
Process the ECCC oil records that were reconciled by Rin.  He merged the data
for most of the records, and has compiled a spreadsheet list of "CC" record IDs
and their associated AD IDs.

I have placed the file in the same folder as this script ("./CC_to_AD.xlsx").

Input: a list of CC related oil entries that were kept.

    - Column 1 ("CC ID"): The CC ID oil entry that is still up on the
                          Stage Server.
    - Column 2 ("New AD ID"): The AD ID that was a match to the CC in the
                              first column.

Chris wants the first column names to be changed to the second column names.
Rin went through and got rid of the other issues (i.e IDs for the CC that we
got rid of or kept because there were no exact matches), so there are only
the CCs that we kept that need to be renamed.
"""
import os
import sys
from pathlib import Path
from argparse import ArgumentParser

from openpyxl import load_workbook

import adios_db
from adios_db.models.oil.oil import Oil


argp = ArgumentParser(description='Script Arguments:')
argp.add_argument('--path', nargs=1,
                  help=('Specify a path to a data storage area (filesystem). '
                        'If not specified, the default is to use "./data"'))
argp.add_argument('-d', '--dry_run', action='store_true',
                  help=('Do not perform the file actions, only print '
                        'the actions out.'))

def generate_sheet(file):
    wb = load_workbook(file, data_only=True)
    return wb['Sheet1']  # First sheet is not set in the file, use default.


def generate_row_iter(sheet):
    sheet_iter = sheet.iter_rows()
    sheet_iter.__next__()  # bypass column names

    return sheet_iter


def oil_json_file_path(base_path, collection_name, oil_id):
    return (Path(base_path) / collection_name / oil_id[:2] / f'{oil_id}.json')


def clobber_file(cc_obj, cc_file_path, ad_file_path, dry_run):
    print(f'clobbering {cc_file_path} into {ad_file_path}')
    if dry_run:
        return

    os.remove(ad_file_path)
    cc_obj.to_file(ad_file_path)
    os.remove(cc_file_path)


def rename_file(cc_obj, cc_file_path, ad_file_path, dry_run):
    print(f'rename {cc_file_path} to {ad_file_path}')
    if dry_run:
        return

    # The AD file doesn't exist, so the underlying paths might not either
    dir_path = os.path.dirname(ad_file_path)
    os.makedirs(dir_path, exist_ok=True)

    cc_obj.to_file(ad_file_path)
    os.remove(cc_file_path)


def set_id_from_filename(oil_obj, oil_file_path, dry_run):
    print(f'set_id_from_filename {oil_file_path}')
    oil_obj.oil_id = oil_file_path.name.split('.')[0]


def diag_print_oil_fields(oil_obj):
    msg = f'''        {oil_obj.oil_id=},
        {oil_obj.metadata.labels=},
        {oil_obj.metadata.alternate_names=},
        {oil_obj.metadata.comments=},
        {oil_obj.metadata.reference.reference=},
        {oil_obj.status=},
        {oil_obj.metadata.gnome_suitable=}
    '''
    print(msg)


def get_fields_from_oil(ad_oil):
    new_reference_content = ('\n\n'
                             'As published in: '
                             'Environment and Climate Change Canada, '
                             'A Catalogue of Crude Oil and Oil Product Properties '
                             '(1999)- Revised 2022, '
                             'Environment and Climate Change Canada, 2022.')

    return (ad_oil.oil_id,
            ad_oil.metadata.labels,
            ad_oil.metadata.alternate_names,
            ad_oil.metadata.comments,
            ad_oil.metadata.reference.reference + new_reference_content)


def update_oil_fields(oil_obj, oil_id,
                      labels, alternate_names, comments, reference):
    oil_obj.oil_id = oil_id
    oil_obj.metadata.labels += labels
    oil_obj.metadata.alternate_names += alternate_names

    if len(comments) > 0:
        # we do have something to add
        if len(oil_obj.metadata.comments.strip()) == 0:
            oil_obj.metadata.comments = comments
        elif oil_obj.metadata.comments.strip().endswith('.'):
            oil_obj.metadata.comments += f'  {comments}'
        else:
            oil_obj.metadata.comments += f', {comments}'

    if len(reference) > 0:
        # we do have something to add
        if len(oil_obj.metadata.reference.reference.strip()) == 0:
            oil_obj.metadata.reference.reference = reference
        elif oil_obj.metadata.comments.strip().endswith('.'):
            oil_obj.metadata.reference.reference += f'\n{reference}'
        else:
            oil_obj.metadata.reference.reference += f',\n{reference}'

    oil_obj.status = None
    oil_obj.metadata.gnome_suitable = None


def set_id_from_filename(oil_obj, oil_file_path, dry_run):
    print(f'set_id_from_filename {oil_file_path}')
    oil_obj.oil_id = oil_file_path.name.split('.')[0]


def diag_print_oil_fields(oil_obj):
    msg = f'''        {oil_obj.oil_id=},
        {oil_obj.metadata.labels=},
        {oil_obj.metadata.alternate_names=},
        {oil_obj.metadata.comments=},
        {oil_obj.metadata.reference.reference=}
    '''
    print(msg)


def get_fields_from_oil(ad_oil):
    new_reference_content = ('\n\n'
                             'As published in: '
                             'Environment and Climate Change Canada, '
                             'A Catalogue of Crude Oil and Oil Product Properties '
                             '(1999)- Revised 2022, '
                             'Environment and Climate Change Canada, 2022.')

    return (ad_oil.oil_id,
            ad_oil.metadata.labels,
            ad_oil.metadata.alternate_names,
            ad_oil.metadata.comments,
            ad_oil.metadata.reference.reference + new_reference_content)


def update_oil_fields(oil_obj, oil_id,
                      labels, alternate_names, comments, reference):
    oil_obj.oil_id = oil_id
    oil_obj.metadata.labels += labels
    oil_obj.metadata.alternate_names += alternate_names

    if len(comments) > 0:
        # we do have something to add
        if len(oil_obj.metadata.comments.strip()) == 0:
            oil_obj.metadata.comments = comments
        elif oil_obj.metadata.comments.strip().endswith('.'):
            oil_obj.metadata.comments += f'  {comments}'
        else:
            oil_obj.metadata.comments += f', {comments}'

    if len(reference) > 0:
        # we do have something to add
        if len(oil_obj.metadata.reference.reference.strip()) == 0:
            oil_obj.metadata.reference.reference = reference
        elif oil_obj.metadata.comments.strip().endswith('.'):
            oil_obj.metadata.reference.reference += f'\n{reference}'
        else:
            oil_obj.metadata.reference.reference += f',\n{reference}'

    oil_obj.status = None
    oil_obj.metadata.gnome_suitable = None


def main(argv=sys.argv):
    args = argp.parse_args(argv[1:])

    base_path = Path(args.path[0]) if args.path is not None else Path('./data')
    print(f'Work directory: {base_path}')

    dry_run = args.dry_run
    if dry_run:
        print(f'Dry run...no files to be changed.')

    sheet = generate_sheet(Path('./CC_to_AD.xlsx'))

    for r in generate_row_iter(sheet):
        cc_id, ad_id = [f.value for f in r]
        cc_file = oil_json_file_path(base_path, "oil", cc_id)
        ad_file = oil_json_file_path(base_path, "oil", ad_id)

        if cc_file.is_file() and ad_file.is_file():
            cc_oil = Oil.from_file(cc_file)
            ad_oil = Oil.from_file(ad_file)

            #print('Before:')
            #diag_print_oil_fields(cc_oil)

            update_oil_fields(cc_oil, *get_fields_from_oil(ad_oil))

            #print('After:')
            #diag_print_oil_fields(cc_oil)

            clobber_file(cc_oil, cc_file, ad_file, dry_run)
        elif cc_file.is_file():
            cc_oil = Oil.from_file(cc_file)

            set_id_from_filename(cc_oil, ad_file, dry_run)
            rename_file(cc_oil, cc_file, ad_file, dry_run)
        else:
            print(f'{cc_file} file is missing.')


if __name__ == "__main__":
    main()
