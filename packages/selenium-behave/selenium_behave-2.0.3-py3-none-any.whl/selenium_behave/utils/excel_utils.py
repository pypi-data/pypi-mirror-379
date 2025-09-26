import openpyxl

excel_path = "Text_To_Voice_With_KeyWord.xlsx"

def get_excel_lines():
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    return [row[0].value for row in sheet.iter_rows(min_row=2) if row[0].value]

def get_common_keywords():
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    keywords = []
    for row in sheet.iter_rows(min_row=2):
        cell = row[1].value
        if cell:
            if "," in str(cell):
                keywords.extend([word.strip().lower() for word in str(cell).split(",")])
            else:
                keywords.append(str(cell).strip().lower())
    return list(set(keywords))
